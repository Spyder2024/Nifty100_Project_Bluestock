"""Excel export with colour-coded scoring for screener results.

Creates ``screener_output.xlsx`` with three sheets:
    1. Summary       – all companies sorted by composite score
    2. By Sector     – same data grouped by sector
    3. Score Details – individual metric-level percentile scores

Score colour coding (quintile-based):
    >= 80  dark green  |  60-79  light green
    40-59  yellow      |  20-39  light red
    <  20  dark red
"""

from pathlib import Path

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Constants ────────────────────────────────────────────────────────────────

SCORE_FILLS = {
    "dark_green": PatternFill(
        start_color="27AE60", end_color="27AE60", fill_type="solid"
    ),
    "light_green": PatternFill(
        start_color="82E0AA", end_color="82E0AA", fill_type="solid"
    ),
    "yellow": PatternFill(
        start_color="F9E79F", end_color="F9E79F", fill_type="solid"
    ),
    "light_red": PatternFill(
        start_color="F1948A", end_color="F1948A", fill_type="solid"
    ),
    "dark_red": PatternFill(
        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
    ),
}

HEADER_FILL = PatternFill(
    start_color="2C3E50", end_color="2C3E50", fill_type="solid"
)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
SCORE_FONT = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Ordered columns for the Summary / By Sector sheets
SUMMARY_COLUMNS = [
    "company_name",
    "sector",
    "composite_score",
    "profitability_score",
    "cash_quality_score",
    "growth_score",
    "leverage_score",
    "return_on_equity",
    "return_on_capital_employed",
    "net_profit_margin",
    "operating_profit_margin",
    "cfo_quality_score",
    "operating_cash_flow_ratio",
    "debt_to_equity",
    "interest_coverage_ratio",
    "revenue_cagr_5yr",
    "net_profit_cagr_5yr",
    "ebitda_cagr_5yr",
]

# Columns that receive quintile colour coding
SCORE_COLUMNS = [
    "composite_score",
    "profitability_score",
    "cash_quality_score",
    "growth_score",
    "leverage_score",
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _score_fill(value) -> PatternFill:
    """Return a PatternFill based on the score quintile (0-100)."""
    if pd.isna(value):
        return PatternFill()          # fill_type=None → no colour
    v = float(value)
    if v >= 80:
        return SCORE_FILLS["dark_green"]
    if v >= 60:
        return SCORE_FILLS["light_green"]
    if v >= 40:
        return SCORE_FILLS["yellow"]
    if v >= 20:
        return SCORE_FILLS["light_red"]
    return SCORE_FILLS["dark_red"]


def _style_header(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 30) -> None:
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value or "")) for c in col_cells),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


def _colour_score_cells(ws, df: pd.DataFrame, score_cols: list[str]) -> None:
    """Apply quintile fills + SCORE_FONT to score columns (rows 2+)."""
    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}
    for row in range(2, len(df) + 2):
        for col_name in score_cols:
            if col_name not in col_map:
                continue
            cell = ws.cell(row=row, column=col_map[col_name])
            cell.fill = _score_fill(cell.value)
            cell.font = SCORE_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER


def _style_body(ws, df: pd.DataFrame, score_cols: list[str]) -> None:
    """Apply basic formatting to non-score cells."""
    score_set = set(score_cols)
    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}
    for row in range(2, len(df) + 2):
        for col_name, col_idx in col_map.items():
            if col_name in score_set:
                continue
            cell = ws.cell(row=row, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_name in ("company_name", "sector") else CENTER


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=1
    ):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


# ── Main Export ───────────────────────────────────────────────────────────────

def export_to_excel(
    df: pd.DataFrame,
    output_path: str,
    composite_scores: pd.DataFrame | None = None,
    sector_col: str = "sector",
    include_details: bool = True,
) -> Path:
    """Export screener results to a colour-coded ``.xlsx`` file.

    Parameters
    ----------
    df : DataFrame
        Screener results.  Must contain ``company_name``, *sector_col*,
        and at least some scoring-metric columns.
    output_path : str
        Destination file path (e.g. ``"download/screener_output.xlsx"``).
    composite_scores : DataFrame, optional
        Pre-computed output of :func:`compute_all_scores`.  If *None*
        the function computes it internally.
    sector_col : str
        Name of the sector column (default ``"sector"``).
    include_details : bool
        Whether to include the *Score Details* sheet (default *True*).

    Returns
    -------
    Path
        Absolute path to the created file.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    ValueError
        If *df* is empty.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel export.  "
            "Install with:  pip install openpyxl"
        )
    if df.empty:
        raise ValueError("Cannot export an empty DataFrame")

    from .scoring import (
        compute_all_scores,
        sector_relative_score,
        ALL_SCORING_METRICS,
    )

    # ── Compute / merge scores ───────────────────────────────────────────
    if composite_scores is None:
        composite_scores = compute_all_scores(df, sector_col=sector_col)

    export_df = df.copy()
    for col in composite_scores.columns:
        if col not in export_df.columns:
            export_df[col] = composite_scores[col]

    # ── Summary DataFrame ────────────────────────────────────────────────
    avail = [c for c in SUMMARY_COLUMNS if c in export_df.columns]
    summary = (
        export_df[avail]
        .sort_values("composite_score", ascending=False, na_position="last")
        .reset_index(drop=True)
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    _write_df_to_sheet(ws1, summary)
    _style_header(ws1, len(avail))
    _colour_score_cells(ws1, summary, SCORE_COLUMNS)
    _style_body(ws1, summary, SCORE_COLUMNS)
    _auto_width(ws1)

    # ── Sheet 2: By Sector ──────────────────────────────────────────────
    ws2 = wb.create_sheet("By Sector")
    by_sector = (
        export_df[avail]
        .sort_values(
            [sector_col, "composite_score"],
            ascending=[True, False],
            na_position="last",
        )
        .reset_index(drop=True)
    )
    _write_df_to_sheet(ws2, by_sector)
    _style_header(ws2, len(avail))
    _colour_score_cells(ws2, by_sector, SCORE_COLUMNS)
    _style_body(ws2, by_sector, SCORE_COLUMNS)
    _auto_width(ws2)

    # ── Sheet 3: Score Details ──────────────────────────────────────────
    if include_details:
        ws3 = wb.create_sheet("Score Details")
        detail_data: dict[str, pd.Series] = {
            "company_name": export_df["company_name"],
            sector_col: export_df[sector_col],
        }
        for metric in ALL_SCORING_METRICS:
            if metric in export_df.columns:
                detail_data[f"{metric}_score"] = sector_relative_score(
                    export_df, metric, sector_col
                ).round(1)
        for sc in SCORE_COLUMNS:
            if sc in export_df.columns:
                detail_data[sc] = export_df[sc]

        detail_df = pd.DataFrame(detail_data)
        detail_df = detail_df.sort_values(
            "composite_score", ascending=False, na_position="last"
        ).reset_index(drop=True)

        _write_df_to_sheet(ws3, detail_df)
        all_score_cols = [c for c in detail_df.columns if c.endswith("_score")]
        _style_header(ws3, len(detail_df.columns))
        _colour_score_cells(ws3, detail_df, all_score_cols)
        _style_body(ws3, detail_df, all_score_cols)
        _auto_width(ws3)

    wb.save(str(out))
    return out.resolve()