"""Peer comparison Excel export (Day 20).

Creates ``peer_comparison.xlsx`` with:
  - 1 **Overview** sheet — one row per peer group with average percentiles
  - 1 sheet per active peer group — companies ranked by composite percentile,
    colour-coded quintile cells, and a highlighted **benchmark row**
    showing the peer-group average.

Colour coding (quintile-based, 0–100):
    >= 80  dark green  |  60–79  light green
    40–59  yellow      |  20–39  light red
    <  20  dark red    |  benchmark  light blue
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .peer import (
    PEER_METRICS,
    load_peer_percentiles,
)

# ── Style constants ────────────────────────────────────────────────────────────

_PCT_FILLS = {
    "dark_green": PatternFill(
        start_color="27AE60", end_color="27AE60", fill_type="solid"
    ),
    "light_green": PatternFill(
        start_color="82E0AA", end_color="82E0AA", fill_type="solid"
    ),
    "yellow": PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid"),
    "light_red": PatternFill(
        start_color="F1948A", end_color="F1948A", fill_type="solid"
    ),
    "dark_red": PatternFill(
        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
    ),
}

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
SCORE_FONT = Font(name="Calibri", size=10, bold=True)
BENCH_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
BENCH_FONT = Font(name="Calibri", bold=True, size=10, color="1A5276")
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ── Small helpers ──────────────────────────────────────────────────────────────


def _pct_fill(value) -> PatternFill:
    """Quintile fill for a 0–100 percentile value."""
    if pd.isna(value):
        return PatternFill()
    v = float(value)
    if v >= 80:
        return _PCT_FILLS["dark_green"]
    if v >= 60:
        return _PCT_FILLS["light_green"]
    if v >= 40:
        return _PCT_FILLS["yellow"]
    if v >= 20:
        return _PCT_FILLS["light_red"]
    return _PCT_FILLS["dark_red"]


def _write_df(ws, df: pd.DataFrame) -> None:
    """Write DataFrame (with header) starting at row 1, col 1."""
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)


def _style_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_w: int = 12, max_w: int = 28) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 3, min_w), max_w
        )


def _get_col_indices(df: pd.DataFrame) -> dict[str, int]:
    return {name: idx + 1 for idx, name in enumerate(df.columns)}


# ── Colour + style application ─────────────────────────────────────────────────


def _colour_score_cells(
    ws,
    df: pd.DataFrame,
    score_cols: list[str],
    start_row: int = 2,
    end_row: int | None = None,
) -> None:
    """Apply quintile fills to score cells in a range."""
    if end_row is None:
        end_row = start_row + len(df) - 1
    col_map = _get_col_indices(df)
    for row in range(start_row, end_row + 1):
        for sc in score_cols:
            if sc not in col_map:
                continue
            cell = ws.cell(row=row, column=col_map[sc])
            cell.fill = _pct_fill(cell.value)
            cell.font = SCORE_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER


def _style_body_cells(
    ws,
    df: pd.DataFrame,
    score_cols: list[str],
    start_row: int = 2,
    end_row: int | None = None,
) -> None:
    """Apply plain body formatting to non-score cells."""
    if end_row is None:
        end_row = start_row + len(df) - 1
    col_map = _get_col_indices(df)
    score_set = set(score_cols)
    for row in range(start_row, end_row + 1):
        for col_name, col_idx in col_map.items():
            if col_name in score_set:
                continue
            cell = ws.cell(row=row, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_name == "company_name" else CENTER


def _style_benchmark_row(
    ws,
    df: pd.DataFrame,
    score_cols: list[str],
    benchmark_row: int,
) -> None:
    """Apply distinct benchmark styling to a single row."""
    col_map = _get_col_indices(df)
    score_set = set(score_cols)
    for col_name, col_idx in col_map.items():
        cell = ws.cell(row=benchmark_row, column=col_idx)
        cell.fill = BENCH_FILL
        cell.font = BENCH_FONT if col_name in score_set else BODY_FONT
        cell.alignment = CENTER if col_name in score_set else LEFT
        cell.border = THIN_BORDER


# ── Sheet builders ─────────────────────────────────────────────────────────────


def _build_overview(
    ws,
    all_data: pd.DataFrame,
    metrics: list[str],
    active_groups: list[str],
) -> None:
    """One row per peer group with average percentiles + overall average."""
    rows: list[dict] = []
    for pg in active_groups:
        pg_data = all_data[all_data["peer_group"] == pg]
        row: dict = {
            "Peer Group": pg,
            "# Companies": pg_data["company_name"].nunique(),
        }
        for m in metrics:
            vals = pg_data.loc[pg_data["metric_name"] == m, "percentile_rank"]
            row[m] = round(vals.mean(), 1) if not vals.empty else None
        row["Overall Avg"] = round(pg_data["percentile_rank"].mean(), 1)
        rows.append(row)

    df = (
        pd.DataFrame(rows)
        .sort_values("Overall Avg", ascending=False)
        .reset_index(drop=True)
    )

    _write_df(ws, df)
    score_cols = [m for m in metrics if m in df.columns] + ["Overall Avg"]
    _style_header(ws, len(df.columns))
    _colour_score_cells(ws, df, score_cols)
    _style_body_cells(ws, df, score_cols)
    _auto_width(ws)


def _build_peer_sheet(
    ws,
    pg_data: pd.DataFrame,
    metrics: list[str],
) -> None:
    """Companies in a peer group ranked by composite, + benchmark row."""
    # Pivot long → wide
    pivot = pg_data.pivot_table(
        index="company_name",
        columns="metric_name",
        values="percentile_rank",
        aggfunc="first",
    ).reset_index()

    # Ensure every metric column exists
    for m in metrics:
        if m not in pivot.columns:
            pivot[m] = None

    metric_cols = [m for m in metrics if m in pivot.columns]

    # Composite = mean of available metric percentiles
    pivot["Composite"] = pivot[metric_cols].mean(axis=1).round(1)
    pivot = pivot.sort_values("Composite", ascending=False).reset_index(drop=True)

    # Benchmark row (peer-group average)
    benchmark: dict = {"company_name": "Peer Group Average"}
    for m in metric_cols:
        benchmark[m] = round(pivot[m].mean(), 1)
    benchmark["Composite"] = round(pivot["Composite"].mean(), 1)
    pd.DataFrame([benchmark])

    cols = ["company_name"] + metric_cols + ["Composite"]
    data_df = pivot[cols].copy()

    # Write data
    _write_df(ws, data_df)

    # Blank separator row
    len(data_df) + 2
    # Write benchmark
    bench_start = len(data_df) + 2
    for c_idx, col in enumerate(cols, start=1):
        ws.cell(row=bench_start, column=c_idx, value=benchmark.get(col))

    score_cols = metric_cols + ["Composite"]
    n_data = len(data_df)

    _style_header(ws, len(cols))
    _colour_score_cells(ws, data_df, score_cols, start_row=2, end_row=1 + n_data)
    _style_body_cells(ws, data_df, score_cols, start_row=2, end_row=1 + n_data)
    _style_benchmark_row(ws, data_df, score_cols, benchmark_row=bench_start)
    _auto_width(ws)


# ── Public API ─────────────────────────────────────────────────────────────────


def export_peer_comparison_excel(
    conn,
    output_path: str,
    year: int,
    metrics: Optional[list[str]] = None,
) -> Path:
    """Export peer comparison data to a multi-sheet ``.xlsx`` file.

    Parameters
    ----------
    conn : sqlite3.Connection
        Database with a populated ``peer_percentiles`` table.
    output_path : str
        Destination file path.
    year : int
        Year to export.
    metrics : list[str], optional
        Metrics to include.  Defaults to all 10 ``PEER_METRICS``.

    Returns
    -------
    Path
        Resolved path to the created file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required.  Install with: pip install openpyxl")

    if metrics is None:
        metrics = PEER_METRICS

    all_data = load_peer_percentiles(conn, year=year)
    if all_data.empty:
        raise ValueError(f"No peer percentile data for year {year}")

    active_groups = sorted(all_data["peer_group"].unique())

    wb = Workbook()

    # ── Overview ──────────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    _build_overview(ws_ov, all_data, metrics, active_groups)

    # ── One sheet per peer group ──────────────────────────────────────────
    for pg in active_groups:
        pg_data = all_data[all_data["peer_group"] == pg]
        # Excel sheet name max 31 chars
        safe_name = pg[:31]
        ws = wb.create_sheet(title=safe_name)
        _build_peer_sheet(ws, pg_data, metrics)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out.resolve()
