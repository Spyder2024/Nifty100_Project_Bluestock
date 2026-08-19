"""
[S5] fix: migrate cash_flow null columns from source XLSX
Root cause: ETL created operating_cf / investing_cf / financing_cf
           but XLSX headers are operating_activity / investing_activity / financing_activity
This script reads data/raw/cashflow.xlsx and backfills the NULL columns.
"""
import argparse
import sqlite3
import sys
from pathlib import Path

# ── resolve project root from script location ──────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from src.etl.normaliser import normalize_ticker, normalize_year

# ── CLI ────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(
    description="Backfill cash_flow NULL columns from cashflow.xlsx"
)
parser.add_argument(
    "--dry-run",
    action="store_true",
    help="Print planned UPDATEs without writing to the DB",
)
ARGS = parser.parse_args()

# ── auto-discover DB (prefer script-anchored, fall back to CWD) ─
DB_CANDIDATES = [
    PROJECT_ROOT / "output" / "nifty100.db",
    PROJECT_ROOT / "data" / "output" / "nifty100.db",
    Path("output/nifty100.db"),
    Path("data/output/nifty100.db"),
    Path("nifty100.db"),
]
DB_PATH = next((p for p in DB_CANDIDATES if p.exists()), None)
if not DB_PATH:
    print(f"[ERROR] nifty100.db not found. Tried: {[str(p) for p in DB_CANDIDATES]}")
    sys.exit(1)

# ── cashflow source (script-anchored first) ──────────────────
CSV_CANDIDATES = [
    PROJECT_ROOT / "data" / "raw" / "cashflow.xlsx",
    PROJECT_ROOT / "data" / "cash_flow.csv",
    PROJECT_ROOT / "cash_flow.csv",
    Path("data/raw/cashflow.xlsx"),
    Path("data/cash_flow.csv"),
    Path("cash_flow.csv"),
    Path("output/cash_flow.csv"),
]
CSV_PATH = next((p for p in CSV_CANDIDATES if p.exists()), None)
if not CSV_PATH:
    print(f"[ERROR] cashflow source not found. Tried: {[str(p) for p in CSV_CANDIDATES]}")
    sys.exit(1)

print(f"[INFO] DB  : {DB_PATH}")
print(f"[INFO] CSV : {CSV_PATH}")

# ── column mapping (CSV name → DB column name) ────────────────
COLUMN_MAP = {
    "operating_activity": "operating_cf",
    "investing_activity": "investing_cf",
    "financing_activity": "financing_cf",
    "net_cash_flow": "net_cash_flow",
}

# ── load cashflow Excel (row 1 = banner, row 2 = real headers) ─
wb = load_workbook(CSV_PATH, data_only=True, read_only=True)
ws = wb.active

# Sanity check the banner row
banner = ws.cell(row=1, column=1).value
if not isinstance(banner, str) or "Cash Flow" not in banner:
    print(f"[ERROR] Unexpected banner row in {CSV_PATH.name}: {banner!r}. Refusing to continue.")
    sys.exit(1)

# Read header row and assert the expected columns
EXPECTED_HEADERS = {
    "company_id", "year",
    "operating_activity", "investing_activity",
    "financing_activity", "net_cash_flow",
}
header = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
missing = EXPECTED_HEADERS - set(header)
if missing:
    print(f"[ERROR] {CSV_PATH.name} is missing expected columns: {sorted(missing)}")
    print(f"[ERROR] Got: {header}")
    sys.exit(1)
col_idx = {name: header.index(name) for name in EXPECTED_HEADERS}

# Iterate data rows (skip banner row 1 and header row 2)
rows_by_key: dict[tuple[str, str], dict[str, float | None]] = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if row is None or all(v is None for v in row):
        continue

    raw_company = row[col_idx["company_id"]] if col_idx["company_id"] < len(row) else None
    raw_year = row[col_idx["year"]] if col_idx["year"] < len(row) else None
    if raw_company is None or raw_year is None:
        continue

    company = normalize_ticker(str(raw_company))
    if company == "AGTL":
        company = "ATGL"
    year = normalize_year(str(raw_year))
    if company is None or year is None:
        continue

    mapped: dict[str, float | None] = {}
    for csv_col, db_col in COLUMN_MAP.items():
        idx = col_idx.get(csv_col)
        if idx is None or idx >= len(row):
            continue
        val = row[idx]
        if val is None:
            continue
        if isinstance(val, str) and val.strip() in ("", "-", "—"):
            continue
        try:
            mapped[db_col] = float(val)
        except (ValueError, TypeError):
            mapped[db_col] = None

    if mapped:
        rows_by_key[(company, year)] = mapped  # last-write-wins on duplicates
wb.close()

print(f"[INFO] Loaded {len(rows_by_key)} (company, year) keys from {CSV_PATH.name}")

# ── connect & update ──────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

# verify cash_flow table exists and has target columns
cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='cash_flow'")
if not cur.fetchone():
    print("[ERROR] cash_flow table not found in DB.")
    sys.exit(1)

cur.execute("PRAGMA table_info(cash_flow)")
db_cols = {r[1] for r in cur.fetchall()}
print(f"[INFO] DB cash_flow columns: {sorted(db_cols)}")

for db_col in ("operating_cf", "investing_cf", "financing_cf"):
    if db_col not in db_cols:
        print(f"[WARN] Column '{db_col}' not in table — will ADD it.")
        cur.execute(f"ALTER TABLE cash_flow ADD COLUMN {db_col} REAL")

# ── fetch existing rows to match ──────────────────────────────
cur.execute("SELECT rowid, company_id, year FROM cash_flow")
db_rows = cur.fetchall()

db_keys: set[tuple[str, str]] = set()
updated = 0
not_found = 0
try:
    for rowid, company, year in db_rows:
        key = (str(company).upper().strip(), str(year).strip())
        db_keys.add(key)
        if key in rows_by_key:
            vals = rows_by_key[key]
            set_clauses = []
            params = []
            for db_col, val in vals.items():
                if val is not None and db_col in db_cols:
                    set_clauses.append(f"{db_col} = ?")
                    params.append(val)
            if set_clauses:
                sql = f"UPDATE cash_flow SET {', '.join(set_clauses)} WHERE rowid = ?"
                params.append(rowid)
                if ARGS.dry_run:
                    print(f"[DRY-RUN] {sql}  params={params}")
                else:
                    cur.execute(sql, params)
                updated += 1
        else:
            not_found += 1
    if not ARGS.dry_run:
        conn.commit()
except Exception:
    conn.rollback()
    raise

# CSV-side rows the DB doesn't have (gap ≈ 1188 XLSX rows - 1056 DB rows)
csv_only = [k for k in rows_by_key if k not in db_keys]

# ── verify ────────────────────────────────────────────────────
print(f"\n[RESULT] Updated         : {updated} rows")
print(f"[RESULT] DB no match     : {not_found} rows")
print(f"[RESULT] CSV no DB match : {len(csv_only)} rows  (e.g. {csv_only[:5]})")

for col in ("operating_cf", "investing_cf", "financing_cf"):
    cur.execute(f"SELECT COUNT(*) FROM cash_flow WHERE {col} IS NOT NULL")
    cnt = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM cash_flow")
    total = cur.fetchone()[0]
    pct = (cnt / total * 100) if total else 0
    print(f"[VERIFY] {col}: {cnt}/{total} non-null ({pct:.1f}%)")

conn.close()
print("[DONE] Migration complete. Re-run Day 31 cashflow_kpis.py now.")