"""Tests for peer comparison Excel export (Day 20)."""

import sqlite3
from pathlib import Path

import pandas as pd
import pytest

from src.analytics.peer import (
    PEER_METRICS,
    create_peer_table,
    save_peer_percentiles,
)
from src.analytics.peer_export import (
    export_peer_comparison_excel,
    _pct_fill,
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _seed_peer_data(conn, year=2024):
    """Populate peer_percentiles with 3 peer groups × 2-3 companies."""
    data = []
    companies = {
        "IT": [
            ("TCS",  {"return_on_equity": 83.3, "net_profit_margin": 83.3, "debt_to_equity": 83.3, "interest_coverage_ratio": 83.3}),
            ("INFY", {"return_on_equity": 50.0, "net_profit_margin": 50.0, "debt_to_equity": 50.0, "interest_coverage_ratio": 50.0}),
        ],
        "Financial Services": [
            ("HDFCBANK", {"return_on_equity": 66.7, "net_profit_margin": 66.7, "debt_to_equity": 66.7, "interest_coverage_ratio": 66.7}),
            ("SBIN",     {"return_on_equity": 33.3, "net_profit_margin": 33.3, "debt_to_equity": 33.3, "interest_coverage_ratio": 33.3}),
        ],
        "FMCG": [
            ("HINDUNILVR", {"return_on_equity": 75.0, "net_profit_margin": 75.0, "debt_to_equity": 75.0, "interest_coverage_ratio": 75.0}),
        ],
    }

    # Only seed a subset of metrics to keep fixture lean
    seed_metrics = ["return_on_equity", "net_profit_margin", "debt_to_equity", "interest_coverage_ratio"]

    for pg, members in companies.items():
        for name, pctls in members:
            for metric, pr in pctls.items():
                if metric in seed_metrics:
                    data.append((name, year, pg, metric, pr, len(members)))

    conn.executemany(
        "INSERT INTO peer_percentiles "
        "(company_name, year, peer_group, metric_name, percentile_rank, peer_count) "
        "VALUES (?,?,?,?,?,?)",
        data,
    )
    conn.commit()
    return seed_metrics


@pytest.fixture
def db_and_metrics():
    conn = sqlite3.connect(":memory:")
    create_peer_table(conn)
    seeded = _seed_peer_data(conn, year=2024)
    yield conn, seeded
    conn.close()


# ── _pct_fill helper ──────────────────────────────────────────────────────────

class TestPctFill:
    def test_dark_green(self):
        assert _pct_fill(85).fill_type == "solid"

    def test_nan_no_fill(self):
        assert _pct_fill(None).fill_type is None

    def test_different_ranges_different_colours(self):
        colours = {_pct_fill(v).start_color.rgb for v in [90, 70, 50, 30, 10]}
        assert len(colours) == 5


# ── Basic export ──────────────────────────────────────────────────────────────

class TestBasicExport:
    def test_creates_file(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        result = export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_has_overview_sheet(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Overview" in wb.sheetnames

    def test_has_one_sheet_per_peer_group(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        # Overview + IT + Financial Services + FMCG = 4
        assert len(wb.sheetnames) == 4
        assert "IT" in wb.sheetnames
        assert "Financial Services" in wb.sheetnames
        assert "FMCG" in wb.sheetnames


# ── Overview sheet ────────────────────────────────────────────────────────────

class TestOverviewSheet:
    def test_has_peer_group_column(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["Overview"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Peer Group" in headers

    def test_has_overall_avg_column(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["Overview"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Overall Avg" in headers

    def test_one_row_per_peer_group(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["Overview"]
        data_rows = ws.max_row - 1  # minus header
        assert data_rows == 3  # IT, Financial Services, FMCG

    def test_sorted_by_overall_avg_desc(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["Overview"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        avg_col = headers.index("Overall Avg") + 1
        avgs = [ws.cell(r, avg_col).value for r in range(2, ws.max_row + 1)]
        assert avgs == sorted(avgs, reverse=True)

    def test_score_cells_have_colour(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["Overview"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        avg_col = headers.index("Overall Avg") + 1
        cell = ws.cell(2, avg_col)
        assert cell.fill.fill_type == "solid"


# ── Peer group sheets ─────────────────────────────────────────────────────────

class TestPeerGroupSheet:
    def test_has_company_name_column(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "company_name" in headers

    def test_has_composite_column(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Composite" in headers

    def test_sorted_by_composite_desc(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        comp_col = headers.index("Composite") + 1
        # Row 2 and 3 are data (TCS=83.3, INFY=50.0)
        assert ws.cell(2, comp_col).value >= ws.cell(3, comp_col).value

    def test_has_benchmark_row(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        # Last data row + 1 = benchmark row
        # 2 data rows (TCS, INFY) → data rows 2-3, benchmark at row 4
        last_row = ws.max_row
        name_cell = ws.cell(last_row, 1)
        assert name_cell.value == "Peer Group Average"

    def test_benchmark_row_styled(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        bench_row = ws.max_row
        # Benchmark row should have distinct fill
        bench_fill = ws.cell(bench_row, 2).fill.start_color.rgb
        data_fill = ws.cell(2, 2).fill.start_color.rgb
        assert bench_fill != data_fill

    def test_benchmark_composite_is_average(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "peer.xlsx"
        export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        comp_col = headers.index("Composite") + 1
        tcs_comp = ws.cell(2, comp_col).value
        infy_comp = ws.cell(3, comp_col).value
        bench_comp = ws.cell(4, comp_col).value
        expected = round((tcs_comp + infy_comp) / 2, 1)
        assert bench_comp == pytest.approx(expected, abs=0.15)


# ── Edge cases ────────────────────────────────────────────────────────────────

class TestEdgeCases:
    def test_no_data_raises(self, tmp_path):
        conn = sqlite3.connect(":memory:")
        create_peer_table(conn)
        out = tmp_path / "empty.xlsx"
        with pytest.raises(ValueError, match="No peer"):
            export_peer_comparison_excel(conn, str(out), 2024)
        conn.close()

    def test_custom_metrics_subset(self, db_and_metrics, tmp_path):
        conn, _ = db_and_metrics
        out = tmp_path / "subset.xlsx"
        export_peer_comparison_excel(
            conn, str(out), 2024, metrics=["return_on_equity", "debt_to_equity"]
        )
        from openpyxl import load_workbook
        ws = load_workbook(str(out))["IT"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # Should only have company_name + the 2 metrics + Composite
        assert len(headers) == 4

    def test_creates_parent_directories(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        nested = tmp_path / "a" / "b" / "out.xlsx"
        export_peer_comparison_excel(conn, str(nested), 2024, metrics=metrics)
        assert nested.exists()

    def test_return_path_is_absolute(self, db_and_metrics, tmp_path):
        conn, metrics = db_and_metrics
        out = tmp_path / "abs.xlsx"
        result = export_peer_comparison_excel(conn, str(out), 2024, metrics=metrics)
        assert result.is_absolute()