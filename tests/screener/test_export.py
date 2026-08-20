"""Tests for Excel export with colour-coded scoring (Day 17)."""

import numpy as np
import pandas as pd
import pytest

from src.screener.scoring import compute_all_scores
from src.screener.export import (
    export_to_excel,
    _score_fill,
    SCORE_COLUMNS,
)

# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture
def sample_df():
    """4 companies × 2 sectors × 11 metrics — minimal export test data."""
    return pd.DataFrame(
        {
            "company_name": ["TCS", "INFY", "HDFCBANK", "SBIN"],
            "sector": ["IT", "IT", "Financial Services", "Financial Services"],
            "return_on_equity": [15.0, 12.0, 14.0, 10.0],
            "return_on_capital_employed": [20.0, 16.0, 0.0, 0.0],
            "net_profit_margin": [18.0, 15.0, 20.0, 12.0],
            "operating_profit_margin": [22.0, 19.0, 25.0, 16.0],
            "cfo_quality_score": [0.85, 0.80, 0.70, 0.55],
            "operating_cash_flow_ratio": [1.2, 1.0, 0.0, 0.0],
            "debt_to_equity": [0.1, 0.3, 5.8, 4.2],
            "interest_coverage_ratio": [25.0, 20.0, 2.5, 1.8],
            "revenue_cagr_5yr": [8.0, 7.0, 10.0, 8.0],
            "net_profit_cagr_5yr": [10.0, 8.0, 12.0, 9.0],
            "ebitda_cagr_5yr": [9.0, 7.5, 11.0, 8.5],
        }
    )


# ── Score-fill helper ────────────────────────────────────────────────────────


class TestScoreFill:
    def test_high_score_solid_fill(self):
        assert _score_fill(85.0).fill_type == "solid"

    def test_nan_no_fill(self):
        assert _score_fill(np.nan).fill_type is None

    def test_different_ranges_different_colours(self):
        fills = {_score_fill(v).start_color.rgb for v in [90, 70, 50, 30, 10]}
        assert len(fills) == 5, "Each quintile must produce a unique colour"

    def test_boundary_80_same_as_100(self):
        assert _score_fill(80.0).start_color.rgb == _score_fill(100.0).start_color.rgb

    def test_boundary_60_same_as_70(self):
        assert _score_fill(60.0).start_color.rgb == _score_fill(70.0).start_color.rgb

    def test_79_diff_from_80(self):
        assert _score_fill(79.0).start_color.rgb != _score_fill(80.0).start_color.rgb

    def test_boundary_40(self):
        assert _score_fill(40.0).start_color.rgb == _score_fill(55.0).start_color.rgb

    def test_boundary_20(self):
        assert _score_fill(20.0).start_color.rgb == _score_fill(35.0).start_color.rgb

    def test_boundary_19_is_dark_red(self):
        assert _score_fill(19.0).start_color.rgb == _score_fill(5.0).start_color.rgb


# ── File-level export tests ──────────────────────────────────────────────────


class TestExportToFile:
    def test_creates_file(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        result = export_to_excel(sample_df, str(out))
        assert result.exists() and result.stat().st_size > 0

    def test_has_summary_sheet(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        assert "Summary" in load_workbook(str(out)).sheetnames

    def test_has_by_sector_sheet(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        assert "By Sector" in load_workbook(str(out)).sheetnames

    def test_has_score_details_by_default(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        assert "Score Details" in load_workbook(str(out)).sheetnames

    def test_no_details_when_disabled(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out), include_details=False)
        from openpyxl import load_workbook

        assert "Score Details" not in load_workbook(str(out)).sheetnames

    def test_summary_sorted_by_composite_desc(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        ws = load_workbook(str(out))["Summary"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ci = headers.index("composite_score") + 1
        scores = [ws.cell(r, ci).value for r in range(2, ws.max_row + 1)]
        assert scores == sorted(scores, reverse=True)

    def test_summary_has_all_score_columns(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        ws = load_workbook(str(out))["Summary"]
        headers = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
        for col in SCORE_COLUMNS:
            assert col in headers, f"{col} missing from Summary"

    def test_header_row_styled(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        ws = load_workbook(str(out))["Summary"]
        cell = ws.cell(1, 1)
        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"

    def test_score_cells_have_colour_fill(self, sample_df, tmp_path):
        out = tmp_path / "test.xlsx"
        export_to_excel(sample_df, str(out))
        from openpyxl import load_workbook

        ws = load_workbook(str(out))["Summary"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ci = headers.index("composite_score") + 1
        cell = ws.cell(2, ci)
        assert cell.fill.fill_type == "solid"


# ── Edge cases ───────────────────────────────────────────────────────────────


class TestExportEdgeCases:
    def test_empty_df_raises_value_error(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        with pytest.raises(ValueError, match="empty"):
            export_to_excel(pd.DataFrame({"company_name": [], "sector": []}), str(out))

    def test_creates_parent_directories(self, sample_df, tmp_path):
        nested = tmp_path / "a" / "b" / "out.xlsx"
        export_to_excel(sample_df, str(nested))
        assert nested.exists()

    def test_pre_computed_scores_used(self, sample_df, tmp_path):
        out = tmp_path / "pre.xlsx"
        pre = compute_all_scores(sample_df)
        export_to_excel(sample_df, str(out), composite_scores=pre)
        from openpyxl import load_workbook

        assert "Summary" in load_workbook(str(out)).sheetnames

    def test_return_path_matches(self, sample_df, tmp_path):
        out = tmp_path / "match.xlsx"
        result = export_to_excel(sample_df, str(out))
        assert result == out.resolve()
